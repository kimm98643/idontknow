import pandas as pd
import openpyxl

def process_eis_to_excel(wb, temp_val, mpt_path, start_line, radius, thickness, drt_target_temp):
    if temp_val in wb.sheetnames: ws = wb[temp_val]
    else: ws = wb.copy_worksheet(wb.worksheets[0]); ws.title = temp_val

    ws['C3'], ws['E5'], ws['C5'] = float(radius), float(thickness) / 2, int(temp_val)
    df = pd.read_csv(mpt_path, sep='\t', skiprows=start_line, encoding='cp1252')
    
    if df.iloc[0, 0] < df.iloc[-1, 0]: df = df.iloc[::-1].reset_index(drop=True)

    drt_out = []
    is_target = str(temp_val).strip() == str(drt_target_temp).strip()

    for i in range(len(df)):
        r = i + 9
        f, re, im = float(df.iloc[i, 0]), float(df.iloc[i, 1]), float(df.iloc[i, 2])
        ws.cell(row=r, column=2, value=f)
        ws.cell(row=r, column=3, value=re)
        ws.cell(row=r, column=4, value=im)
        if is_target:
            # 원본 데이터 그대로 전달 (여기서 나누지 않음!)
            drt_out.append([f, re, im])

    last_row, ref_row = len(df) + 8, 9
    for r in range(last_row, 8, -1):
        try:
            val = ws.cell(row=r, column=4).value
            if val is not None and float(val) < 0:
                ref_row = r; break
        except: continue
    for r in range(9, last_row + 1):
        if ws.cell(row=r, column=2).value:
            ws.cell(row=r, column=9, value=f"=G{r}-$G${ref_row}")
            ws.cell(row=r, column=10, value=f"=H{r}")

    return wb, drt_out

def finalize_drt_results(wb, drt_temp, tau_list, gamma_list):
    ws = wb[drt_temp]
    # D3 셀에서 면적 계수 읽기 (수식 대신 값 직접 삽입 → 메모리 절약)
    try:
        area_coeff = float(ws['D3'].value) if ws['D3'].value else 1.0
    except (TypeError, ValueError):
        area_coeff = 1.0

    for i in range(len(tau_list)):
        r = i + 9
        tau_val = tau_list[i]
        gamma_val = gamma_list[i]
        ws.cell(row=r, column=14, value=1.0 / tau_val)  # f = 1/τ (τ는 1/f convention)
        ws.cell(row=r, column=15, value=gamma_val * area_coeff)                     # gamma × area
        ws.cell(row=r, column=12, value=f"=B{r}")
        ws.cell(row=r, column=13, value=f"=J{r}")
    return wb

def create_summary_sheet(wb, sorted_temps):
    if "Summary" not in wb.sheetnames: ws = wb.create_sheet("Summary", 0)
    else: ws = wb["Summary"]
    for i, t in enumerate(sorted_temps, start=2):
        ws.cell(row=i, column=2, value=int(t))
        ws.cell(row=i, column=3, value=f"='{t}'!J5")
        ws.cell(row=i, column=4, value=f"=1000/(B{i}+273.15)")
        ws.cell(row=i, column=5, value=f"=LN(C{i})")