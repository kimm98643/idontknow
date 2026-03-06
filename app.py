import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import os, re, io, gc, tempfile, traceback
from scipy.optimize import nnls

# NumPy 2.0+ 호환: trapz → trapezoid
_trapz = getattr(np, 'trapezoid', None) or np.trapz

# ──────────────────────────────────────────────
#  1. EIS 로더 (eis_loader.py 통합)
# ──────────────────────────────────────────────

def find_target_files_from_uploads(uploaded_files):
    """업로드된 .mpt 파일에서 온도 추출."""
    pattern = re.compile(r'_(\d{3,4})C?_C')
    temp_map = {}   # { "700": UploadedFile, ... }
    temps = []
    for uf in uploaded_files:
        m = pattern.search(uf.name)
        if m:
            t = m.group(1).strip()
            if t.endswith(('00', '50')) and t not in temps:
                temps.append(t)
                temp_map[t] = uf
    return temp_map, temps


def get_data_start_line(file_bytes):
    """바이트에서 'freq/Hz' 행을 찾는다."""
    try:
        text = file_bytes.decode('cp1252')
        for i, line in enumerate(text.splitlines()):
            if 'freq/Hz' in line:
                return i
    except Exception:
        pass
    return 65


# ──────────────────────────────────────────────
#  2. DRT 솔버 (drt_solver.py 통합)
# ──────────────────────────────────────────────

def solve_drt_core(freq, z_re, z_im, mode, lam):
    freq = np.asarray(freq, dtype=float).copy()
    z_re = np.asarray(z_re, dtype=float).copy()
    z_im = np.asarray(z_im, dtype=float).copy()

    # 부호 안전 검사: z_im은 -Im(Z) (양수)여야 함
    if np.mean(z_im) < 0:
        z_im = -z_im

    if mode == 3:
        mask = z_im >= 0
        freq, z_re, z_im = freq[mask], z_re[mask], z_im[mask]

    omega = 2 * np.pi * freq
    n = len(freq)
    tau_pts = 1.0 / omega

    si = np.argsort(tau_pts)
    tau_pts, omega = tau_pts[si], omega[si]
    z_re, z_im = z_re[si], z_im[si]
    log_tau = np.log(tau_pts)

    coeff = 0.5
    dlt = np.abs(np.mean(np.diff(log_tau))) if n > 1 else 1.0
    eps = 2.0 * np.sqrt(np.log(2.0)) / (coeff * dlt)

    margin = 5.0 / eps
    NI = 1500
    x = np.linspace(np.min(log_tau) - margin, np.max(log_tau) + margin, NI)
    dx = x[1] - x[0]

    we = omega[:, None] * np.exp(x)[None, :]
    dn = 1.0 + we ** 2
    re_k = 1.0 / dn;  im_k = we / dn
    del we, dn

    A_re = np.zeros((n, n));  A_im = np.zeros((n, n))
    dc = np.empty((n, NI))
    for k in range(n):
        d = x - log_tau[k]
        r = np.exp(-(eps * d) ** 2)
        dc[k] = -2.0 * eps ** 2 * d * r
        A_re[:, k] = _trapz(re_k * r[None, :], dx=dx, axis=1)
        A_im[:, k] = _trapz(im_k * r[None, :], dx=dx, axis=1)
    del re_k, im_k

    M = np.zeros((n, n))
    for k in range(n):
        M[k, k:] = _trapz(dc[k] * dc[k:], dx=dx, axis=1)
        M[k:, k] = M[k, k:]
    del dc;  gc.collect()

    ev, evc = np.linalg.eigh(M)
    U = np.diag(np.sqrt(np.maximum(ev, 0.0))) @ evc.T
    del ev, evc, M

    if mode == 2:
        nv = n + 2
        Kr = np.zeros((n, nv));  Kr[:, 1] = 1.0;  Kr[:, 2:] = A_re
        Ki = np.zeros((n, nv));  Ki[:, 0] = -omega; Ki[:, 2:] = A_im
        Kg = np.zeros((n, nv));  Kg[:, 2:] = np.sqrt(lam) * U
    else:
        nv = n + 1
        Kr = np.zeros((n, nv));  Kr[:, 0] = 1.0;  Kr[:, 1:] = A_re
        Ki = np.zeros((n, nv));  Ki[:, 1:] = A_im
        Kg = np.zeros((n, nv));  Kg[:, 1:] = np.sqrt(lam) * U
    del A_re, A_im, U

    Ka = np.vstack((Kr, Ki, Kg))
    Za = np.concatenate((z_re, z_im, np.zeros(n)))
    del Kr, Ki, Kg;  gc.collect()

    xo, _ = nnls(Ka, Za);  del Ka, Za
    gw = xo[2:] if mode == 2 else xo[1:]

    NF = 500
    tmx = np.ceil(np.max(np.log10(1.0 / freq))) + 0.5
    tmn = np.floor(np.min(np.log10(1.0 / freq))) - 0.5
    ff = np.logspace(-tmn, -tmx, NF)
    te = 1.0 / (2.0 * np.pi * ff)

    gf = np.zeros(NF)
    for m in range(n):
        if gw[m] == 0.0: continue
        gf += gw[m] * np.exp(-(eps * np.log(te / tau_pts[m])) ** 2)

    gc.collect()
    return (1.0 / ff).tolist(), gf.tolist()


# ──────────────────────────────────────────────
#  3. Excel 처리 (excel_processor.py 통합)
# ──────────────────────────────────────────────

def process_eis_to_excel(wb, temp_val, file_bytes, start_line, radius, thickness, drt_target_temp):
    if temp_val in wb.sheetnames:
        ws = wb[temp_val]
    else:
        ws = wb.copy_worksheet(wb.worksheets[0]);  ws.title = temp_val

    ws['C3'] = float(radius)
    ws['E5'] = float(thickness) / 2
    ws['C5'] = int(temp_val)

    text = file_bytes.decode('cp1252')
    lines = text.splitlines()
    header_idx = start_line
    data_lines = lines[header_idx + 1:]

    rows = []
    for line in data_lines:
        parts = line.strip().split('\t')
        if len(parts) >= 3:
            try:
                rows.append([float(parts[0]), float(parts[1]), float(parts[2])])
            except ValueError:
                continue
    if not rows:
        return wb, []

    df = pd.DataFrame(rows, columns=['freq', 're', 'im'])
    if df.iloc[0, 0] < df.iloc[-1, 0]:
        df = df.iloc[::-1].reset_index(drop=True)

    drt_out = []
    is_target = str(temp_val).strip() == str(drt_target_temp).strip()

    for i in range(len(df)):
        r = i + 9
        f, re_v, im_v = float(df.iloc[i, 0]), float(df.iloc[i, 1]), float(df.iloc[i, 2])
        ws.cell(row=r, column=2, value=f)
        ws.cell(row=r, column=3, value=re_v)
        ws.cell(row=r, column=4, value=im_v)
        if is_target:
            drt_out.append([f, re_v, im_v])

    last_row = len(df) + 8
    ref_row = 9
    for r in range(last_row, 8, -1):
        try:
            val = ws.cell(row=r, column=4).value
            if val is not None and float(val) < 0:
                ref_row = r;  break
        except:
            continue
    for r in range(9, last_row + 1):
        if ws.cell(row=r, column=2).value:
            ws.cell(row=r, column=9, value=f"=G{r}-$G${ref_row}")
            ws.cell(row=r, column=10, value=f"=H{r}")

    return wb, drt_out


def finalize_drt_results(wb, drt_temp, tau_list, gamma_list):
    if drt_temp not in wb.sheetnames:
        raise KeyError(f"시트 '{drt_temp}'를 찾을 수 없습니다. 온도 이름을 확인하세요.")
    ws = wb[drt_temp]
    try:
        area_coeff = float(ws['D3'].value) if ws['D3'].value else 1.0
    except (TypeError, ValueError):
        area_coeff = 1.0

    for i in range(len(tau_list)):
        r = i + 9
        tau_val = float(tau_list[i])
        gamma_val = float(gamma_list[i])
        # N열: τ → freq 변환 (1/τ). τ=0이면 안전하게 0
        if tau_val != 0:
            ws.cell(row=r, column=14, value=1.0 / tau_val)
        else:
            ws.cell(row=r, column=14, value=0.0)
        # O열: γ × 면적계수
        ws.cell(row=r, column=15, value=gamma_val * area_coeff)
        ws.cell(row=r, column=12, value=f"=B{r}")
        ws.cell(row=r, column=13, value=f"=J{r}")
    return wb


def create_summary_sheet(wb, sorted_temps):
    if "Summary" not in wb.sheetnames:
        ws = wb.create_sheet("Summary", 0)
    else:
        ws = wb["Summary"]
    for i, t in enumerate(sorted_temps, start=2):
        ws.cell(row=i, column=2, value=int(t))
        ws.cell(row=i, column=3, value=f"='{t}'!J5")
        ws.cell(row=i, column=4, value=f"=1000/(B{i}+273.15)")
        ws.cell(row=i, column=5, value=f"=LN(C{i})")


# ══════════════════════════════════════════════
#  4. Streamlit UI
# ══════════════════════════════════════════════

st.set_page_config(page_title="EIS → DRT Analyzer", layout="wide")
st.title("⚡ EIS → DRT 분석기")
st.caption("MATLAB DRTtools와 동일한 계산 결과 · Streamlit Cloud 배포용")

# ── 사이드바: 설정 ──
with st.sidebar:
    st.header("⚙️ 설정")
    radius = st.number_input("반지름 (cm)", value=0.5, step=0.01, format="%.3f")
    thickness = st.number_input("두께 (cm)", value=0.1, step=0.001, format="%.4f")
    drt_target_temp = st.text_input("DRT 분석 온도", value="700")
    output_name = st.text_input("출력 파일 이름", value="결과")

    st.divider()
    st.header("📂 템플릿 파일")
    tmpl_file = st.file_uploader("엑셀 템플릿 (.xlsm)", type=["xlsm", "xlsx"])

# ── 메인: 파일 업로드 ──
st.header("1️⃣ EIS 데이터 업로드")
uploaded_files = st.file_uploader(
    ".mpt 파일을 업로드하세요 (여러 개 가능)",
    type=["mpt"], accept_multiple_files=True
)

if uploaded_files and tmpl_file:
  try:
    # 온도 추출
    file_map, temps = find_target_files_from_uploads(uploaded_files)
    sorted_temps = sorted(temps, key=int, reverse=True)

    if not sorted_temps:
        st.error("❌ 파일명에서 온도를 추출할 수 없습니다. '_숫자_C' 패턴이 필요합니다.")
        st.stop()

    st.success(f"✅ {len(sorted_temps)}개 온도 감지: {', '.join(sorted_temps)}°C")

    # 템플릿 로드
    tmpl_bytes = tmpl_file.read()
    tmpl_file.seek(0)
    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp:
        tmp.write(tmpl_bytes)
        tmp_path = tmp.name
    wb = openpyxl.load_workbook(tmp_path, keep_vba=True)

    # Excel 변환
    drt_input = []
    progress = st.progress(0, text="Excel 변환 중...")
    for i, t in enumerate(sorted_temps):
        uf = file_map[t]
        fbytes = uf.read();  uf.seek(0)
        sl = get_data_start_line(fbytes)
        wb, current_drt = process_eis_to_excel(
            wb, t, fbytes, sl, radius, thickness, drt_target_temp
        )
        if current_drt:
            drt_input = current_drt
        progress.progress((i + 1) / len(sorted_temps), text=f"{t}°C 처리 완료")

    if not drt_input:
        st.error(f"❌ '{drt_target_temp}'°C 데이터를 찾지 못했습니다.")
        st.stop()

    st.header("2️⃣ DRT 분석")
    d = np.array(drt_input)
    freq_raw = d[:, 0]
    re_half = d[:, 1] / 2
    # .mpt 3열 = -Im(Z) (양수). 만약 Im(Z)(음수)로 저장된 파일이면 자동 보정
    im_col = d[:, 2]
    if np.mean(im_col) < 0:
        im_col = -im_col          # Im(Z) → -Im(Z) 변환
    neg_im_half = im_col / 2      # 대칭셀 보정: -Im(Z)/2

    # ── 15개 그래프 계산 ──
    lambdas = [1e-1, 1e-2, 1e-3, 1e-4, 1e-5]
    modes = [1, 2, 3]
    mode_names = ["w/o Induc", "with Induc", "Discard"]

    all_results = []
    calc_progress = st.progress(0, text="DRT 계산 중...")
    total = len(modes) * len(lambdas)
    cnt = 0
    for mode in modes:
        for lam in lambdas:
            tau, gamma = solve_drt_core(freq_raw, re_half, neg_im_half, mode, lam)
            all_results.append((tau, gamma))
            cnt += 1
            calc_progress.progress(cnt / total, text=f"계산 중... ({cnt}/{total})")

    calc_progress.empty()
    st.success("✅ 15개 DRT 계산 완료!")

    # ── 그래프 표시 ──
    st.header("3️⃣ 그래프 선택")
    fig, axes = plt.subplots(3, 5, figsize=(22, 12))
    idx = 0
    for mi, mode in enumerate(modes):
        for li, lam in enumerate(lambdas):
            tau, gamma = all_results[idx]
            ax = axes[mi, li]
            ax.semilogx(tau, gamma, color='navy', linewidth=2.0)
            ax.set_title(f"#{idx+1}\n{mode_names[mi]}, λ={lam:.0e}", fontsize=10)
            ax.set_xlabel("τ (s)", fontsize=8)
            ax.set_ylabel("γ(τ)", fontsize=8)
            ax.grid(True, alpha=0.3)
            idx += 1
    plt.tight_layout()
    st.pyplot(fig)
    plt.close(fig)

    # ── 사용자 선택 ──
    labels = []
    idx = 0
    for mi, mode in enumerate(modes):
        for li, lam in enumerate(lambdas):
            labels.append(f"#{idx+1} — {mode_names[mi]}, λ={lam:.0e}")
            idx += 1

    choice = st.selectbox("마음에 드는 그래프를 선택하세요:", labels)
    choice_idx = labels.index(choice)

    # 선택된 그래프 확대
    sel_tau, sel_gamma = all_results[choice_idx]
    fig2, ax2 = plt.subplots(figsize=(10, 5))
    ax2.semilogx(sel_tau, sel_gamma, color='navy', linewidth=2.5)
    ax2.set_xlabel("τ (s)", fontsize=14)
    ax2.set_ylabel("γ(τ) (Ω)", fontsize=14)
    ax2.set_title(f"선택: {labels[choice_idx]}", fontsize=14)
    ax2.grid(True, alpha=0.3)
    plt.tight_layout()
    st.pyplot(fig2)
    plt.close(fig2)

    # ── 4. 저장 ──
    st.header("4️⃣ 결과 저장")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("📊 Excel 저장", type="primary", use_container_width=True):
            with st.spinner("Excel 생성 중..."):
                wb = finalize_drt_results(wb, drt_target_temp, sel_tau, sel_gamma)
                create_summary_sheet(wb, sorted_temps)

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

            st.download_button(
                label="⬇️ Excel 다운로드",
                data=buf,
                file_name=f"{output_name}.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                use_container_width=True,
            )

    with col2:
        # DRT 텍스트 파일 (MATLAB export 형식과 동일)
        txt_lines = ["tau, gamma(tau)\n"]
        for t_val, g_val in zip(sel_tau, sel_gamma):
            txt_lines.append(f"{t_val:.6e}, {g_val:.6e}\n")
        txt_data = "".join(txt_lines)

        st.download_button(
            label="⬇️ DRT 텍스트 다운로드",
            data=txt_data,
            file_name=f"{output_name}_DRT.txt",
            mime="text/plain",
            use_container_width=True,
        )

    # 클린업
    try:
        os.remove(tmp_path)
    except Exception:
        pass
    gc.collect()

  except Exception as e:
    st.error("❌ 오류 발생!")
    st.code(traceback.format_exc(), language="python")
    st.info(f"오류 타입: {type(e).__name__}\n메시지: {e}")

elif not tmpl_file:
    st.info("👈 사이드바에서 엑셀 템플릿 파일을 먼저 업로드하세요.")
else:
    st.info("☝️ .mpt 파일을 업로드하세요.")
